from openpyxl import Workbook
from openpyxl.styles import Font

## Get the data

exec(open('dataBuild.py').read())

## Define what data is important for the tabs

headers = {
    'all': { #Headers for both media and series, the lists is [series_order,medias_order]
        # These [1,1] and etc lists are a failed attempt at making this code more dynamic and I have not bothered to fix it yet
    'Name': [1,1],
    'Artist(s)':[2,2],
    'Link': [4,4],
    'HQ Source(s)': [5,7],
    'Alternate Names': [6,6]
    },
    'media': { #Headers only for medias Numbers are order of the collumns
    'Related To': 5,
    'Order': 3,
    'Earliest Date': 8
},
    'series': { #Headers only for series (all medias tab)
        'Order': 3,
        'Earliest Date': 8,
        'Medias': 7
    }
}
extra_info = [
    'truename',
    'sourcelinks'
]

series_order = { #to-do: Merge these two order dicts with the one above later
    1 : 'Name',
    2 : 'Artist(s)',
    3 : 'Order',
    4 : 'Link',
    5 : 'HQ Source(s)',
    6 : 'Alternate Names',
    7 : 'Medias',
    8 : 'Earliest Date'
}

tabs_order = {
    1 : 'Name',
    2 : 'Artist(s)',
    3: 'Order',
    4: 'Link',
    5: 'Related To',
    6: 'Alternate Names',
    7 : 'HQ Source(s)',
    8 : 'Earliest Date'
}

## Format data into tab styles

tabs = {'series': {}}

for x in medias: #medias is from dataBuild.py
    tabs[x] = {}

for x in headers['all']:
    for y in tabs:
        tabs[y][x] = []
for x in headers['media']:
    for y in medias:
        tabs[y][x] = []
for x in headers['series']:
    tabs['series'][x] = []
for x in extra_info:
    tabs['series'][x] = []
    for y in medias:
        tabs[y][x] = []

ordering = {} #To fix order later
for x in tabs:
    ordering[x] = {}

# Iterate through every song to populate the tabs

exchanging_all = {'Name' : 'name', 'Artist(s)': 'artists', "Link": 'link', "HQ Source(s)": 'sources', 'Alternate Names': 'altnames'}

def DateStrToVal(x):
    val = ''
    for y in x:
        if y != '/':
            val += y
    return float(val)

def ProperDate(x):
    return x[0:10]

def DateNumber(x):
    date_number = DateStrToVal(ProperDate(x))
    if len(x) > 10:
        new_number = float(x[11:])
    else:
        new_number = 99
    return date_number * 100 + new_number

for song in all_data: #all_data is form dataBuild.py
    present_medias = []
    dates = []
    for media in medias:
        try:
            mediadict = song[media]
            present_medias.append(media) # Finding all medias present
            usages = mediadict['uses'] # Find "Related To"
            tabs[media]['Related To'].append(usages) #Append Related To
            thisdate = ProperDate(mediadict['date']) #Find "Date" for the media
            truedate = mediadict['truedate']
            if truedate == 1: # Check whether date is hypothetical
                tabs[media]['Earliest Date'].append(thisdate) #Append first use's date to Earliest Date
            else:
                tabs[media]['Earliest Date'].append('?') #Default Unknown date
            dates.append([mediadict['date'], truedate]) #For later
        except:
            try:
                dummy = song[media]
            except:
                pass
     # Getting the sources and sourcelinks from the medias   
    versions = song['versions']
    sources = [] # Storing the strings
    sourcelinks = []
    for x in versions:
        sources.append(x['source'])
        try:# In case there is no sourcelinks or source
            sourcelinks.append(x['sourcelinks'])
        except:
            pass
    sources = list(set(sources)) #Removing duplicates
    sourcelinks = list(set(sourcelinks))
    source = '' # Strings that will contain the sources
    sourcelink = ''
    for x in sources: #Making them neat strings
        source += ' + ' + x
    source = source[3:]
    for x in sourcelinks:
        sourcelink += ',' + x
    sourcelink = sourcelink[1:]
    song['sources'] = source
    if sourcelink != '': #Only add if there is a link
        song['sourcelinks'] = sourcelink
    for media in present_medias:
        ordering[media][song['name']] = DateNumber(song[media]['date'])
    for x in exchanging_all:
        try:
            tabs['series'][x].append(song[exchanging_all[x]])
        except:
            tabs['series'][x].append(None)
        for media in present_medias:
            try:
                tabs[media][x].append(song[exchanging_all[x]])
            except:
                tabs[media][x].append(None)
    nummericaldates = []
    for value in dates: #Defining Earliest Dates
        nummericaldates.append(DateStrToVal(ProperDate(value[0])))
    lowestvalue = 30000000
    lowpos = 0
    i = 0
    for x in nummericaldates:
        if type(x) == type(1.0):
            if x < lowestvalue:
                lowestvalue = x
                lowpos = i
        i += 1
    try:
        if dates[lowpos][1] == 1:
            tabs['series']['Earliest Date'].append(ProperDate(dates[lowpos][0]))
        else:
            tabs['series']['Earliest Date'].append('?')
    except:
        tabs['series']['Earliest Date'].append('?')
    present_medias = [medias[l]['media'] for l in present_medias]
    media_str = ''
    for x in present_medias:
        media_str += ', ' + x
    tabs['series']['Medias'].append(media_str[2:len(media_str)])
    ordering['series'][song['name']] = DateNumber(dates[lowpos][0])
    for x in extra_info:
        try:
            tabs['series'][x].append(song[x])
        except:
            tabs['series'][x].append(None)
        for y in medias:
            try:
                dummy = song[y]
                try:
                    tabs[y][x].append(song[x])
                except:
                    tabs[y][x].append(None)
            except:
                pass


# Fix the non intenger orders

neworder = {}

for x in ordering:
    neworder[x] = {}
    ordering[x] = dict(sorted(ordering[x].items(), key=lambda z:z[1]))
    i = 0
    for y in ordering[x]:
        i += 1
        neworder[x][y] = i
    for y in tabs[x]['Name']:
        tabs[x]['Order'].append(neworder[x][y])

# Sort tabs by order

orderedtabs = {}

for x in tabs:
    orderedtabs[x] = {}
    for y in tabs[x]:
        orderedtabs[x][y] = []
    for i in range(1,len(tabs[x]['Order'])+1):
        songno = tabs[x]['Order'].index(i)
        for y in orderedtabs[x]:
            orderedtabs[x][y].append(tabs[x][y][songno])

### Creating the spreadsheet

wb = Workbook()

# Configure base sheets

wb['Sheet'].title = 'Welcome'
wb.create_sheet('Downloads')

#Create Series OST's

for k in medias:
    wb.create_sheet(medias[k]['tab'])

#Extra sheets

cpseries = 'Series OST'

wb.create_sheet(cpseries)
wb.create_sheet('to-do')

#Fill data to media sheets

letters = {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J'}

for k in medias:
    tab = medias[k]['tab']
    for i in range(1, len(tabs_order)+1):
        column = orderedtabs[k][tabs_order[i]]
        for j in range(1,len(column)+1):
            cell = letters[i] + str(j)
            if tabs_order[i] == 'Name':
                wb[tab][cell].value = column[j-1]
                if orderedtabs[k]['truename'][j-1] == 1:
                    wb[tab][cell].font = Font(color='2E47AA')
                else:
                    wb[tab][cell].font = Font(color='CC0000')
            elif tabs_order[i] == 'HQ Source(s)':
                wb[tab][cell] = column[j-1]
                if orderedtabs[k]['sourcelinks'][j-1] != None:
                    wb[tab][cell].hyperlink = orderedtabs[k]['sourcelinks'][j-1]
                    wb[tab][cell].style = "Hyperlink"
            elif tabs_order[i] == 'Link':
                if column[j-1] != "" and column[j-1] != None:
                    wb[tab][cell].hyperlink = column[j-1]
                    wb[tab][cell].value = 'Link'
                    wb[tab][cell].style = "Hyperlink"
            else:
                wb[tab][cell] = column[j-1]

#Fill data to series sheet
k = 'series'
tab = cpseries
for i in range(1, len(series_order)+1):
    column = orderedtabs[k][series_order[i]]
    for j in range(1,len(column)+1):
        cell = letters[i] + str(j)
        if series_order[i] == 'Name':
            wb[tab][cell].value = column[j-1]
            if orderedtabs[k]['truename'][j-1] == 1:
                wb[tab][cell].font = Font(color='2E47AA')
            else:
                wb[tab][cell].font = Font(color='CC0000')
        elif series_order[i] == 'HQ Source(s)':
            wb[tab][cell] = column[j-1]
            if orderedtabs[k]['sourcelinks'][j-1] != None:
                wb[tab][cell].hyperlink = orderedtabs[k]['sourcelinks'][j-1]
                wb[tab][cell].style = "Hyperlink"
        elif series_order[i] == 'Link':
            if column[j-1] != "" and column[j-1] != None:
                wb[tab][cell].hyperlink = column[j-1]
                wb[tab][cell].value = 'Link'
                wb[tab][cell].style = "Hyperlink"
        else:
            wb[tab][cell] = column[j-1]

def dateToNumber(x):
    if x == "?":
        return False
    year = 0
    month = 0
    day = 0
    foundYear = False
    foundMonth = False
    for char in x:
        if char == "/":
            if foundYear:
                foundMonth = True
            else:
                foundYear = True
        else:
            if foundYear:
                if foundMonth:
                    day = day * 10 + int(char)
                else:
                    month = month * 10 + int(char)
            else:
                year = year * 10 + int(char)
    return day + month * 100 + year * 10000
        
wb.save('sheet.xlsx')
